"""시나리오 파일(.xlsx/.csv) → list[dict] 로더.

헤더 행을 자동 탐지하고, column_mapper 로 표준 4필드(id/scenario/query/checkpoint)를
유연하게 매핑한다. 매핑 결과는 config.COL_* 에 반영된다.
"""
import csv
from pathlib import Path

import config
import column_mapper as cm


def _rows_from(path):
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"시나리오 파일 없음: {path}")
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    with open(path, encoding="utf-8-sig") as f:
        return [list(r) for r in csv.reader(f)]


def _detect_header(rows):
    """앞쪽 15행 중 '시나리오 헤더다움' 점수가 가장 높은 행을 헤더로 선택."""
    best_idx, best_score, best_mapping, best_header = 0, -1, None, []
    extra = getattr(config, "COLUMN_ALIASES_EXTRA", None)
    for i, row in enumerate(rows[:15]):
        cells = [(str(c).strip() if c is not None else "") for c in row]
        if not any(cells):
            continue
        score, mapping = cm.header_quality(cells, extra)
        if score > best_score:
            best_idx, best_score, best_mapping, best_header = i, score, mapping, cells
    return best_idx, best_header, best_mapping


def load_cases(path=None, verbose=True):
    path = path or config.SCENARIO_PATH
    rows = _rows_from(path)
    if not rows:
        return []

    hidx, header, mapping = _detect_header(rows)
    extra = getattr(config, "COLUMN_ALIASES_EXTRA", None)
    mapping = cm.resolve(header, extra)   # 필수 누락 시 에러
    config.set_columns(mapping)

    if verbose:
        print("[헤더 매핑] " + " | ".join(
            f"{f}→'{mapping.get(f)}'" for f in ["id", "scenario", "query", "checkpoint"]
        ))

    out = []
    for r in rows[hidx + 1:]:
        cells = [(str(c).strip() if c is not None else "") for c in r] if r else []
        if not any(cells):
            continue
        rec = {}
        for i in range(len(header)):
            if not header[i]:
                continue
            rec[header[i]] = cells[i] if i < len(cells) else ""
        if rec.get(config.COL_ID):
            out.append(rec)
    return out
