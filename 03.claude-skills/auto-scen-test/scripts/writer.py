"""테스트 결과를 결과 XLSX(`단위테스트결과` 시트)에 자동 입력 + 스크린샷 임베드."""
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

import config

RESULT_SHEET = "단위테스트결과"
SUMMARY_SHEET = "결과 요약"
RESULT_COLS = ["테스트결과", "실행일시", "응답 요약", "증적 이미지", "검증자/비고"]


def update_xlsx(results):
    """results: [{tc_id, verdict, elapsed, response_text, note, executed_at}, ...]"""
    if not config.XLSX_PATH.exists():
        print(f"[!] {config.XLSX_PATH} 미존재. build_result_xlsx.py 로 먼저 생성하세요.")
        return

    wb = load_workbook(config.XLSX_PATH)
    ws = wb[RESULT_SHEET] if RESULT_SHEET in wb.sheetnames else wb.worksheets[0]

    header = [c.value for c in ws[1]]

    # 결과 컬럼이 없으면 끝에 추가
    for c in RESULT_COLS:
        if c not in header:
            ws.cell(row=1, column=len(header) + 1, value=c)
            header.append(c)

    def col(name):
        return header.index(name) + 1

    col_id = col(config.COL_ID)
    col_verdict = col("테스트결과")
    col_executed = col("실행일시")
    col_summary = col("응답 요약")
    col_image = col("증적 이미지")
    col_note = col("검증자/비고")

    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_id).value
        if v:
            id_to_row[str(v)] = r

    updated = embedded = 0
    for res in results:
        row = id_to_row.get(str(res["tc_id"]))
        if not row:
            continue
        ws.cell(row=row, column=col_verdict, value=res["verdict"])
        ws.cell(row=row, column=col_executed, value=res["executed_at"])
        ws.cell(row=row, column=col_summary, value=res["response_text"])
        ws.cell(row=row, column=col_note, value=res["note"])

        shot = config.SHOTS_DIR / f"{res['tc_id']}.png"
        if shot.exists():
            try:
                img = XLImage(str(shot))
                img.width, img.height = 160, 100
                ws.add_image(img, f"{get_column_letter(col_image)}{row}")
                embedded += 1
            except Exception as e:
                ws.cell(row=row, column=col_image, value=f"[이미지 임베드 실패: {e}]")
        updated += 1

    _update_summary_sheet(wb, ws)
    wb.save(config.XLSX_PATH)
    print(f"[OK] XLSX 갱신: {updated}건 결과 입력, {embedded}건 이미지 임베드 + 요약 갱신")


def _update_summary_sheet(wb, result_ws):
    from collections import Counter
    import re
    from openpyxl.styles import Alignment, Font, PatternFill

    if SUMMARY_SHEET not in wb.sheetnames:
        wb.create_sheet(SUMMARY_SHEET)

    header = [c.value for c in result_ws[1]]
    rc = header.index("테스트결과") + 1
    sid_col = header.index(config.COL_SCENARIO) + 1 if config.COL_SCENARIO in header else None

    verdict_count = Counter()
    scenario_pass = Counter(); scenario_fail = Counter(); scenario_total = Counter()
    scenarios_all = set()
    for r in range(2, result_ws.max_row + 1):
        v = result_ws.cell(row=r, column=rc).value
        sid = result_ws.cell(row=r, column=sid_col).value if sid_col else None
        if sid:
            scenarios_all.add(sid)
            scenario_total[sid] += 1
        if v:
            verdict_count[v] += 1
            if sid and v == "Pass":
                scenario_pass[sid] += 1
            elif sid and v == "Fail":
                scenario_fail[sid] += 1

    total = result_ws.max_row - 1
    pct = lambda n: f"{round(n/total*100, 1)}%" if total else "0%"
    skey = lambda s: int(re.match(r"S(\d+)", str(s)).group(1)) if s and re.match(r"S(\d+)", str(s)) else 999

    ws2 = wb[SUMMARY_SHEET]
    if ws2.max_row:
        ws2.delete_rows(1, ws2.max_row)

    ws2.append(["구분", "건수", "비율"])
    ws2.append(["전체", total, "100%"])
    for v in ["Pass", "Fail", "N/A", "Block", "재시험"]:
        cnt = verdict_count.get(v, 0)
        ws2.append([v, cnt, pct(cnt)])
    미수행 = total - sum(verdict_count.get(v, 0) for v in ["Pass", "Fail", "N/A", "Block", "재시험"])
    ws2.append(["미수행", 미수행, pct(미수행)])
    ws2.append([])
    ws2.append(["시나리오별 진행 현황"])
    ws2.append(["시나리오", "전체", "Pass", "Fail", "미수행"])
    for s in sorted(scenarios_all, key=skey):
        t = scenario_total[s]; p = scenario_pass[s]; f = scenario_fail[s]
        ws2.append([s, t, p, f, t - p - f])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    if ws2.max_row >= 11:
        for cell in ws2[10]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
        for cell in ws2[11]:
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="DDEBF7")
    for c in range(1, 6):
        ws2.column_dimensions[get_column_letter(c)].width = 14
