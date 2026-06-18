"""시나리오 파일로부터 결과 XLSX(`단위테스트결과` + `결과 요약`)를 생성한다.

결과 파일이 아직 없을 때만 사용:
  python build_result_xlsx.py --scenario-file S.xlsx --result-file R.xlsx
"""
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config
from scenario_loader import load_cases

RESULT_COLS = ["테스트결과", "실행일시", "응답 요약", "증적 이미지", "검증자/비고"]


def build(scenario_file, result_file):
    config.SCENARIO_PATH = Path(scenario_file).expanduser().resolve()
    cases = load_cases()
    if not cases:
        raise SystemExit("시나리오에서 케이스를 읽지 못했습니다.")

    # 시나리오 컬럼 순서 유지
    src_cols = list(cases[0].keys())
    out_cols = src_cols + [c for c in RESULT_COLS if c not in src_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "단위테스트결과"
    ws.append(out_cols)
    for c in cases:
        ws.append([c.get(col, "") for col in out_cols])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i in range(1, len(out_cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    wb.create_sheet("결과 요약")

    out = Path(result_file).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] 결과 XLSX 생성: {out} ({len(cases)}건)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--scenario-file", required=True)
    ap.add_argument("--result-file", required=True)
    args = ap.parse_args()
    build(args.scenario_file, args.result_file)
