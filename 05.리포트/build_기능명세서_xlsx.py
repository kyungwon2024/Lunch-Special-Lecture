"""기능명세서.md → Excel 워크북 (6 sheets) 변환기 · 점심특강"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/billy/자동화_경원/Lunch-Special-Lecture/05.리포트/기능명세서_20260531_1551.xlsx"

# ============ 스타일 정의 ============
FONT_HEAD = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Malgun Gothic", size=14, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Malgun Gothic", size=10, color="111827")
FONT_BOLD = Font(name="Malgun Gothic", size=10, bold=True, color="111827")
FONT_META = Font(name="Malgun Gothic", size=9, italic=True, color="6B7280")

FILL_NAVY = PatternFill("solid", start_color="1F2937")
FILL_ORANGE = PatternFill("solid", start_color="FF8F00")
FILL_GRAY_LIGHT = PatternFill("solid", start_color="F3F4F6")
FILL_WHITE = PatternFill("solid", start_color="FFFFFF")
FILL_MUST = PatternFill("solid", start_color="FFE0E0")     # 연한 빨강
FILL_SHOULD = PatternFill("solid", start_color="FFF3E0")   # 연한 주황
FILL_COULD = PatternFill("solid", start_color="E3F2FD")    # 연한 파랑

THIN = Side(border_style="thin", color="D1D5DB")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_HEAD = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CELL = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header_row(ws, row_idx, fill=FILL_NAVY, font=FONT_HEAD):
    for cell in ws[row_idx]:
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_HEAD
        cell.border = BORDER_ALL

def style_data_cells(ws, start_row, end_row, start_col=1, end_col=None):
    end_col = end_col or ws.max_column
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.font or cell.font.name != "Malgun Gothic":
                cell.font = FONT_BODY
            cell.alignment = ALIGN_CELL
            cell.border = BORDER_ALL

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title(ws, title, span_cols):
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=span_cols)
    ws.cell(1, 1).font = FONT_TITLE
    ws.cell(1, 1).fill = FILL_ORANGE
    ws.cell(1, 1).alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

def priority_fill(p):
    if p == "Must" or p == "P0":
        return FILL_MUST
    elif p == "Should" or p == "P1":
        return FILL_SHOULD
    elif p == "Could" or p == "P2":
        return FILL_COULD
    return FILL_WHITE

# ============ 데이터 ============
MUST_LIST = [
    ("F-0101", "위치 기반 지도 뷰", "REQ-F1-001", "Must", "M2 알파"),
    ("F-0102", "리스트 뷰 + 토글", "REQ-F1-002", "Must", "M2 알파"),
    ("F-0103", "반경 필터 (500m/1km/2km)", "REQ-F1-003", "Must", "M2 알파"),
    ("F-0104", "카테고리 필터", "REQ-F1-004", "Must", "M2 알파"),
    ("F-0105", "1인 좌석 필터", "REQ-F1-005", "Must", "M2 알파"),
    ("F-0106", "가격대 필터", "REQ-F1-006", "Must", "M2 알파"),
    ("F-0107", "\"바로 입장 가능\" 필터/정렬", "REQ-F1-007", "Must", "M3 베타"),
    ("F-0108", "매장·메뉴 검색", "REQ-F1-008", "Must", "M2 알파"),
    ("F-0201", "\"오늘의 점심특선\" 메인 피드", "REQ-F2-001", "Must", "M3 베타"),
    ("F-0202", "점심특선 카드 (4요소)", "REQ-F2-002", "Must", "M3 베타"),
    ("F-0203", "실시간 영업 상태 표시", "REQ-F2-003", "Must", "M3 베타"),
    ("F-0204", "예상 식사 소요 시간 메타", "REQ-F2-004", "Must", "M3 베타"),
    ("F-0205", "메인 피드 정렬 옵션", "REQ-F2-005", "Must", "M3 베타"),
    ("F-0206", "점심특선 상세 화면", "REQ-F2-006", "Must", "M3 베타"),
    ("F-0207", "무한 스크롤 페이지네이션", "REQ-F2-007", "Must", "M3 베타"),
    ("F-0301", "쿠폰 발급", "REQ-F3-001", "Must", "M3 베타"),
    ("F-0302", "사용자 본인 인증 (KISA)", "REQ-F3-002", "Must", "M3 베타"),
    ("F-0303", "QR/바코드 토큰 생성 (1분 만료)", "REQ-F3-003", "Must", "M3 베타"),
    ("F-0304", "사장님 QR 스캔 처리", "REQ-F3-004", "Must", "M3 베타"),
    ("F-0305", "위치 기반 사용 검증 (100m)", "REQ-F3-005", "Must", "M3 베타"),
    ("F-0306", "1일 사용 횟수 제한", "REQ-F3-006", "Must", "M3 베타"),
    ("F-0307", "쿠폰 만료 관리", "REQ-F3-007", "Must", "M3 베타"),
    ("F-0308", "쿠폰 사용/만료/취소 알림", "REQ-F3-008", "Must", "M3 베타"),
    ("F-0309", "\"내 쿠폰함\" 화면", "REQ-F3-009", "Must", "M3 베타"),
    ("F-0401", "사업자 인증 (국세청 진위)", "REQ-F4-001", "Must", "M2 알파"),
    ("F-0402", "매장 등록 (3단계 위저드)", "REQ-F4-002", "Must", "M2 알파"),
    ("F-0403", "매장 메타데이터 입력 (좌석·소요시간)", "REQ-F4-003", "Must", "M2 알파"),
    ("F-0404", "일별 점심특선 등록", "REQ-F4-004", "Must", "M3 베타"),
    ("F-0405", "요일별 점심특선 반복 등록", "REQ-F4-005", "Must", "M3 베타"),
    ("F-0406", "즉시/예약 발행", "REQ-F4-006", "Must", "M3 베타"),
    ("F-0407", "실시간 영업 상태 토글", "REQ-F4-007", "Must", "M3 베타"),
    ("F-0408", "매장 정보·메뉴 수정 (이력 30일)", "REQ-F4-008", "Must", "M2 알파"),
    ("F-0501", "쿠폰 발행 (정률·정액·수량·기간)", "REQ-F5-001", "Must", "M3 베타"),
    ("F-0502", "쿠폰 타게팅 (신규/단골/전체)", "REQ-F5-002", "Must", "M3 베타"),
    ("F-0503", "한정 수량 쿠폰 (잔여 카운터)", "REQ-F5-003", "Must", "M3 베타"),
    ("F-0504", "핵심 4지표 대시보드", "REQ-F5-004", "Must", "M3 베타"),
    ("F-0505", "일/주/월 통계 토글", "REQ-F5-005", "Must", "M3 베타"),
    ("F-0506", "ROI 카드 (비용→유입→매출)", "REQ-F5-006", "Must", "M3 베타"),
    ("F-0507", "쿠폰별 성과 분석 + 가이드", "REQ-F5-007", "Must", "M3 베타"),
    ("F-9001", "사용자 회원가입 (소셜 + 본인 인증)", "REQ-OPS-001", "Must", "M2 알파"),
    ("F-9002", "사장님 회원가입 (사업자 인증 흐름)", "REQ-OPS-002", "Must", "M2 알파"),
    ("F-9003", "3단계 권한 분리 (사용자/사장님/어드민)", "REQ-OPS-003", "Must", "M2 알파"),
    ("F-9004", "신고·이의제기 처리 (어드민)", "REQ-OPS-004", "Must", "M3 베타"),
    ("F-9005", "회원 탈퇴 (개인정보 파기·5년 보관)", "REQ-OPS-005", "Must", "M3 베타"),
    ("F-9006", "인앱 알림 센터", "REQ-OPS-006", "Must", "M3 베타"),
    ("F-9007", "큰 글씨 모드 토글", "REQ-OPS-007", "Must", "M3 베타"),
    ("F-9008", "시각약자 접근성 (WCAG 2.1 AA)", "REQ-OPS-008", "Must", "M3 베타"),
]

SHOULD_LIST = [
    ("F-0508", "통계 CSV/Excel 내보내기", "REQ-F5-008", "Should", "M4+"),
    ("F-7001", "사용자 리뷰·별점 작성", "REQ-P2-001", "Should", "M4+"),
    ("F-7002", "즐겨찾기 매장 등록", "REQ-P2-002", "Should", "M4+"),
    ("F-7003", "위치+시간 트리거 푸시 알림", "REQ-P2-003", "Should", "M4+"),
    ("F-7004", "즐겨찾기 신규 특선 푸시", "REQ-P2-004", "Should", "M4+"),
    ("F-7005", "B2B 임직원 식권 자동 발행", "REQ-P2-005", "Should", "M4+"),
    ("F-7006", "B2B 정산 대시보드", "REQ-P2-006", "Should", "M4+"),
    ("F-7007", "저녁·주말 특선 카테고리", "REQ-P2-007", "Should", "M4+"),
]

COULD_LIST = [
    ("F-8001", "광고 상품 정식 운영", "REQ-P3-001", "Could", "M7+"),
    ("F-8002", "토스페이먼츠 선결제", "REQ-P3-002", "Could", "M7+"),
    ("F-8003", "카카오·네이버페이", "REQ-P3-003", "Could", "M7+"),
    ("F-8004", "친구 초대·공유", "REQ-P3-004", "Could", "M7+"),
]

# 풀 명세 (F-ID, 설명, 입력, 처리, 출력, 예외처리, 관련화면, 관련API)
DETAILED = {
    "F-0101": ("사용자의 현재 GPS 위치를 중심으로 주변 식당을 지도 마커로 시각화한다.",
        "userLocation: {lat, lng} (GPS 권한 허용), radius: 500|1000|2000 (디폴트 500m), category[] (선택), mealTimeWindow (현재 시각 11:00~14:00 여부)",
        "① GPS 좌표 획득 (HTML5 Geolocation API) → ② PostGIS ST_DWithin 쿼리로 반경 내 식당 조회 → ③ 카카오맵 SDK에 마커 클러스터링 (10개 이상 시 그룹화) → ④ 점심특선 발행 중인 매장은 오렌지 핀, 일반 매장은 회색 핀",
        "지도 화면 + N개 매장 마커 + 사용자 현재 위치 핀 + 반경 원",
        "❶ GPS 권한 거부 → 수동 위치 검색 폴백 / ❷ 위치 측정 실패 (3초 타임아웃) → 안내 메시지 / ❸ 결과 0건 → 반경 확대 권유 토스트",
        "SCR-USR-001", "API-GET-restaurants?lat=&lng=&radius=&category[]="),
    "F-0102": ("지도 뷰와 동일한 데이터셋을 리스트 형태로 노출. 상단 우측 토글 버튼으로 즉시 전환.",
        "F-0101과 동일 + viewMode: 'map'|'list'",
        "① 동일 API 결과 캐싱 → ② viewMode 토글 시 클라이언트 사이드 렌더 전환 (재요청 없음) → ③ 리스트는 거리 오름차순 기본 정렬",
        "카드형 리스트 (사진/이름/거리/카테고리/점심특선 유무 뱃지)",
        "빈 결과 시 \"주변에 등록된 매장이 없습니다\" + 반경 확대 CTA",
        "SCR-USR-002", "API-GET-restaurants (F-0101과 공유)"),
    "F-0103": ("사용자가 탐색 반경을 3단계로 조절.",
        "radius: 500|1000|2000",
        "반경 변경 시 ① API 재요청 → ② 지도 반경 원 갱신 → ③ 마커/리스트 1초 이내 갱신",
        "갱신된 매장 N개 + 반경 원",
        "결과 0건 → 자동으로 다음 반경 제안 (예: 500m → \"1km로 확대?\")",
        "SCR-USR-001, SCR-USR-002", "API-GET-restaurants"),
    "F-0104": ("한식·분식·카페·양식·중식·일식·건강식/샐러드/단백질 도시락 다중 선택.",
        "category[]: string[]",
        "① 다중 선택 토글 → ② URL 쿼리 동기화 (공유 가능) → ③ API 재요청",
        "선택 카테고리만 노출된 마커/리스트",
        "모든 카테고리 미선택 시 = 전체 노출",
        "SCR-USR-001/002 + SCR-USR-003", "API-GET-restaurants?category[]="),
    "F-0105": ("혼밥형 페르소나를 위한 \"1인 좌석 가능 여부\" 단일 필터.",
        "soloSeating: boolean",
        "식당 메타데이터 solo_seating = true 조건으로 즉시 필터",
        "1인 좌석 매장만 노출",
        "빈 결과 → \"1인 좌석 매장이 적어요. 반경을 확대?\" 친근 메시지",
        "SCR-USR-001/002", "API-GET-restaurants?solo_seating=true"),
    "F-0106": ("점심특선 할인가 기준 4구간(~8천/8천~1만/1만~1.5만/1.5만+) 필터.",
        "priceRange: 'lt8k'|'8k-10k'|'10k-15k'|'gt15k'",
        "점심특선 발행 매장의 할인가 기준 필터링 (정상가 아님)",
        "선택 구간 내 매장만 노출",
        "가격대 0건 → \"다른 가격대도?\" 인접 구간 추천",
        "SCR-USR-001/002", "API-GET-restaurants?priceRange="),
    "F-0107": ("편의점 대비 차별화 — 대기 없이 즉시 입장 가능한 매장만 노출/우선 정렬.",
        "availableNow: boolean",
        "① 사장님 자가 신고 토글(status=available) 또는 ② 운영시간 기반 자동 판정 결합",
        "\"즉시 입장 가능\" 뱃지가 붙은 매장",
        "사장님 미신고 매장은 운영시간 기준만 사용",
        "SCR-USR-001/002", "API-GET-restaurants?available_now=true"),
    "F-0108": ("식당명 또는 메뉴명 부분 일치 검색. 한글 자모 분리 매칭 지원.",
        "q: string (1자 이상)",
        "① 입력 300ms 디바운스 → ② PostgreSQL tsvector + 한글 자모 분리(hangul-js 등) → ③ 결과 20건",
        "매칭된 매장/메뉴 리스트",
        "검색 결과 0건 → \"다른 키워드를 시도하세요\"",
        "SCR-USR-004", "API-GET-search?q="),
    "F-0201": ("사용자 위치 기반 점심특선 카드를 메인 화면에 노출. 11:00~14:00 매장 우선.",
        "userLocation, 현재 시각, sortBy (디폴트: 거리순), 페이지",
        "① 현재 시각 11~14시면 운영 매장 우선 정렬 → ② 그 외 시간엔 \"오늘 점심특선 등록 매장\" 정렬 → ③ 시간 외 매장은 별도 그룹",
        "점심특선 카드 리스트 (20건/페이지)",
        "결과 0건 → \"주변에 점심특선이 없어요. 반경 확대 또는 카테고리 변경\"",
        "SCR-USR-005", "API-GET-lunch-specials?lat=&lng=&radius=&page="),
    "F-0202": ("카드 1개 = 1결정 원칙. 대표 사진/메뉴명/정상가·할인가/할인율% 4요소 노출.",
        "카드 데이터 (서버 응답)",
        "① 할인율 자동 계산 (정상가-할인가)/정상가*100 반올림 → ② 사진은 4:3 비율 lazy load",
        "카드 UI (대표 사진 60% / 정보 40%)",
        "사진 누락 시 카테고리 기본 일러스트",
        "SCR-USR-005", "API-GET-lunch-specials (포함)"),
    "F-0203": ("\"영업중 / 곧 마감 / 휴무\" 3단계 뱃지를 카드 좌상단에 노출.",
        "매장 운영시간 메타 + 사장님 임시 토글",
        "① 운영시간 기준 자동 판정 → ② 사장님이 \"잠시 중단\" 토글 시 즉시 반영 (1분 이내)",
        "색상 뱃지 (녹색/주황/회색)",
        "데이터 불일치 시 운영시간 메타를 우선 신뢰",
        "SCR-USR-005", "API-GET-lunch-specials (포함)"),
    "F-0204": ("15~25분 / 25~40분 / 40분+ 3단계 아이콘으로 카드 우측 하단 표시.",
        "매장 메타 meal_duration",
        "매장 등록 시 필수 입력값을 카드에 그대로 표시",
        "시계 아이콘 + \"15~25분\" 텍스트",
        "미입력 매장은 표시 생략 (등록 단계에서 차단)",
        "SCR-USR-005", "API-GET-lunch-specials (포함)"),
    "F-0205": ("거리순 / 할인율순 / 인기순 / \"바로 입장 가능\"순 4종 정렬.",
        "sortBy: 'distance'|'discount'|'popularity'|'available'",
        "클라이언트 사이드 재정렬 (캐시된 결과 활용) → 1초 이내",
        "재정렬된 카드 리스트",
        "인기순 데이터 없음 시 거리순 폴백",
        "SCR-USR-005", "(클라이언트)"),
    "F-0206": ("메뉴 사진 갤러리 + 사장님 설명 + 잔여 시간 + 위치 지도 + \"쿠폰 받기\" 하단 고정 CTA.",
        "lunchSpecialId",
        "① 카드 탭 시 0.5초 이내 진입 → ② 잔여 수량/시간 실시간 카운터 → ③ 카카오맵 미니맵 임베드",
        "풀스크린 상세 화면",
        "매진 시 CTA 비활성 + \"마감되었습니다\" / 만료 시 자동 비노출",
        "SCR-USR-006", "API-GET-lunch-specials/{id}"),
    "F-0207": ("한 번에 20건 로드, 스크롤 하단 도달 시 다음 페이지 자동 로드.",
        "page: number",
        "① IntersectionObserver로 스크롤 감지 → ② 로딩 스피너 → ③ 가상 리스트로 메모리 관리",
        "추가 20건 append",
        "마지막 페이지 시 \"더 이상 결과가 없어요\" / 네트워크 실패 시 재시도 버튼",
        "SCR-USR-005", "API-GET-lunch-specials?page="),
    "F-0301": ("사용자가 점심특선 상세에서 \"쿠폰 받기\" 탭 시 즉시 발급.",
        "lunchSpecialId, userId (인증 토큰)",
        "① 본인 인증 여부 확인 → ② 1매장/1일 발급 제한 검증 → ③ 한정 수량 잔여 검증 → ④ DB INSERT + Redis 카운터 차감",
        "발급된 쿠폰 객체 + \"내 쿠폰함\" 자동 이동 안내",
        "❶ 미인증 → 본인 인증 흐름으로 이동 / ❷ 일일 제한 초과 → \"내일 다시 시도\" / ❸ 한정 수량 소진 → \"마감\"",
        "SCR-USR-006, SCR-USR-007", "API-POST-coupons/issue"),
    "F-0302": ("KISA 인증 사업자(NICE, KCB 등) SMS 본인 인증. 최초 가입 시 1회.",
        "휴대폰 번호, 인증번호",
        "① 인증 SDK 호출 → ② 결과 토큰 서버 검증 → ③ user.verified = true + 인증 시각 기록",
        "본인 인증 완료 → 원래 흐름 복귀",
        "❶ 인증 시간 초과 (5분) → 재요청 / ❷ 미성년자 → 정책상 차단 / ❸ 동일 휴대폰 다중 계정 → 정지",
        "SCR-USR-008", "API-POST-auth/verify-identity"),
    "F-0303": ("매장 제시용 단기 만료 QR 코드. 스크린샷 도용 방지.",
        "couponId (인증 사용자만)",
        "① JWT 페이로드: {couponId, userId, exp: now+60s} HS256 서명 → ② QR 이미지 생성 → ③ 화면 밝기 자동 최대화",
        "QR 화면 + 잔여 시간 카운터",
        "❶ 만료 → 자동 재발급 (사용자 액션 불필요) / ❷ 토큰 검증 실패 → \"다시 시도\"",
        "SCR-USR-009", "API-POST-coupons/{id}/token"),
    "F-0304": ("사장님 앱 카메라로 QR 스캔 → 사용 처리.",
        "QR에서 추출한 JWT 토큰, 사장님 매장 ID",
        "① 토큰 서명 검증 → ② 만료/사용 여부 확인 → ③ 위치 검증 (F-0305) → ④ coupon.used_at 기록 → ⑤ 사장님/사용자 양쪽 알림",
        "\"사용 완료\" 화면 + 할인 금액 표시",
        "❶ 만료 토큰 / ❷ 이미 사용 / ❸ 위치 불일치 → 각각 안내",
        "SCR-OWN-001", "API-POST-coupons/redeem"),
    "F-0305": ("쿠폰 발급 매장 좌표 ↔ 사용자 GPS 좌표 100m 이내인지 검증.",
        "userLocation, storeLocation",
        "PostGIS ST_Distance 또는 Haversine 계산 → 100m 초과 시 차단",
        "(성공) F-0304 처리 / (실패) 차단 메시지",
        "GPS 권한 없음 → \"위치 확인을 위해 권한이 필요합니다\"",
        "SCR-OWN-001", "API-POST-coupons/redeem (내장)"),
    "F-0306": ("사용자당 일 5건, 매장당 사용자 1건 제한. 양도·도용 방지.",
        "userId, storeId, 날짜",
        "Redis 카운터로 일 단위 사용 횟수 집계 → 초과 시 차단",
        "사용 가능 / 차단",
        "차단 시 \"오늘 사용 한도(5건) 도달\" / \"본 매장 이미 사용\"",
        "SCR-USR-009 / SCR-OWN-001", "(내장 로직)"),
    "F-0307": ("당일 만료 / 7일 만료 / 사장님 지정 만료 3종. 만료 1시간 전 푸시.",
        "쿠폰 발행 시 expiresAt",
        "① 크론잡(5분 주기)으로 만료 임박 쿠폰 식별 → ② FCM 푸시 (옵트인) → ③ 만료 시 자동 status=expired",
        "푸시 알림 + 쿠폰함 자동 비활성",
        "옵트아웃 사용자는 푸시 생략 (인앱 알림은 유지)",
        "SCR-USR-007", "API-GET-coupons/me"),
    "F-0308": ("사용/만료/취소 이벤트 시 사용자·사장님 양쪽에 인앱 + 푸시 알림.",
        "이벤트 트리거 (F-0304/F-0307)",
        "① 알림 큐(BullMQ 등)에 enqueue → ② FCM/APNs 전송 → ③ 인앱 알림 DB INSERT",
        "알림 수신",
        "푸시 토큰 무효 → DB에서 토큰 삭제",
        "SCR-USR-010", "API-GET-notifications"),
    "F-0309": ("사용 가능 / 사용 완료 / 만료 3탭 구조.",
        "userId, 탭",
        "DB 조회 + 만료 임박순 정렬",
        "탭별 쿠폰 리스트",
        "빈 탭 → \"아직 발급받은 쿠폰이 없어요\" + 메인 피드 CTA",
        "SCR-USR-007", "API-GET-coupons/me?status="),
    "F-0401": ("휴대폰 본인 인증 + 사업자 등록번호 입력 + 국세청 API 진위 확인.",
        "휴대폰, 사업자 등록번호, 대표자명",
        "① KISA 본인 인증 (F-0302 재사용) → ② 국세청 진위 확인 API 호출 → ③ owner.verified = true",
        "인증 완료 (평균 5분 이내)",
        "❶ 진위 확인 실패 / ❷ 폐업 사업자 → 각각 안내",
        "SCR-OWN-002", "API-POST-owners/verify-business"),
    "F-0402": ("(1) 기본정보 → (2) 운영시간·메뉴 → (3) 사진·확인. 평균 5분 완료.",
        "단계별 폼",
        "① 단계별 임시 저장 (drafts 테이블) → ② 3단계 완료 후 publish → ③ 운영자 검수 대기 (선택)",
        "매장 ID + 노출 시작",
        "중도 이탈 시 24시간 내 재진입 시 이어쓰기",
        "SCR-OWN-003", "API-POST-stores (draft + publish)"),
    "F-0403": ("solo_seating, seats, meal_duration 등 페르소나 차별화 메타.",
        "Boolean / Integer / Enum",
        "매장 등록 2단계에서 필수 입력 강제",
        "식당 메타 저장",
        "미입력 시 다음 단계 진행 불가",
        "SCR-OWN-003 (2단계)", "API-PATCH-stores/{id}/meta"),
    "F-0404": ("메뉴명, 정상가, 할인가(또는 할인율), 한정 표시, 사진. 30초 내 완료.",
        "폼",
        "① 할인율 자동 계산 → ② 사진 압축·업로드 (Supabase Storage) → ③ 즉시 사용자 피드 노출",
        "점심특선 ID",
        "사진 5MB 초과 시 자동 리사이즈 / 할인가 > 정상가 시 입력 차단",
        "SCR-OWN-004", "API-POST-stores/{id}/lunch-specials"),
    "F-0405": ("매주 월·수 11~14시 등 반복 패턴 설정.",
        "요일[], 시간대, 메뉴 정보",
        "크론 스케줄러로 매일 새벽에 당일 노출 매장 활성화",
        "반복 스케줄 ID",
        "휴무일·임시 휴업 토글 시 해당 일자 자동 제외",
        "SCR-OWN-004", "API-POST-stores/{id}/recurring-specials"),
    "F-0406": ("\"지금 발행\" 또는 \"예약 발행 (분 단위)\". 발행 후 사용 전 한정으로 취소 가능.",
        "publishAt: ISO datetime|now, cancel: boolean",
        "예약 시 큐에 enqueue, 발행 시각 도달 시 활성 처리",
        "발행 예정/완료 상태",
        "이미 사용 시작된 특선은 취소 불가",
        "SCR-OWN-004", "API-PATCH-lunch-specials/{id}/status"),
    "F-0407": ("\"영업중 / 잠시 중단 / 마감\" 3종 토글. \"잠시 중단\"은 30분 후 자동 복귀.",
        "status: 'open'|'paused'|'closed'",
        "① 즉시 사용자 피드 반영 → ② \"paused\"는 30분 타이머 후 \"open\" 자동 전환",
        "매장 상태",
        "운영시간 외에 \"open\" 설정 시 경고 메시지 (강제 차단 X)",
        "SCR-OWN-005", "API-PATCH-stores/{id}/status"),
    "F-0408": ("사진·가격·운영시간 등 수정. 수정 이력 30일 보관 (분쟁 대비).",
        "수정 폼",
        "① 변경 전후 diff 기록 → ② audit log 테이블 INSERT",
        "수정 완료",
        "동시 수정 충돌 시 마지막 저장 우선 + 안내 메시지",
        "SCR-OWN-006", "API-PATCH-stores/{id} + API-GET-stores/{id}/audit"),
    "F-0501": ("정률(%) 또는 정액(원), 적용 메뉴, 수량, 유효기간 설정.",
        "discountType, value, appliedMenus[], quantity, expiresAt",
        "① 권장 할인율 30~40% 가이드 표시 (라스트오더 벤치마크) → ② DB INSERT + 사용자 피드 즉시 노출",
        "쿠폰 ID + 발행 완료",
        "100% 할인 등 비정상 입력 시 경고",
        "SCR-OWN-007", "API-POST-stores/{id}/coupons"),
    "F-0502": ("신규(본 매장 첫 방문) / 단골(직전 30일 N회 이상) / 전체 중 선택.",
        "targetType: 'new'|'loyal'|'all', loyalThreshold",
        "① 사용자 쿠폰 발급 시점에 매장 방문 이력 조회 → ② 조건 부합 사용자에게만 노출",
        "타깃 사용자만 카드 노출",
        "임계값 0 입력 시 디폴트(3회) 적용",
        "SCR-OWN-007", "API-POST-stores/{id}/coupons (포함)"),
    "F-0503": ("발행 수량(예: 50장) 설정 + 잔여 카운터 실시간 노출.",
        "quantity: number",
        "Redis 카운터로 발급마다 차감, 0 도달 시 자동 비노출. 잔여 5장 이하 시 \"마감 임박\" 강조",
        "잔여 수량 표시",
        "동시 발급 race condition → Redis DECR atomic 보장",
        "SCR-OWN-007, SCR-USR-006", "API-POST-coupons/issue (내장)"),
    "F-0504": ("신규 고객 / 30일 재방문율 / 쿠폰 사용률 / 평균 객단가 4카드 1화면.",
        "storeId, period (일/주/월)",
        "배치(시간 단위) 집계 + 실시간 조회 결합",
        "4지표 + 추세 표시 (전 기간 대비 ▲/▼ %)",
        "데이터 부족 (가입 7일 미만) → \"데이터 누적 중\" 안내",
        "SCR-OWN-008", "API-GET-stores/{id}/metrics?period="),
    "F-0505": ("디폴트 주. 토글 1초 이내 전환.",
        "period",
        "클라이언트 사이드 캐싱 + period별 API 호출",
        "갱신된 4지표",
        "일별 최근 30일, 주별 12주, 월별 12개월로 제한",
        "SCR-OWN-008", "API-GET-stores/{id}/metrics"),
    "F-0506": ("\"이번 주 발행 비용 N원 → 신규 N명 → 추정 매출 N원\" — 사장님 락인 핵심.",
        "쿠폰 발행 비용, 사용 매출 데이터",
        "① 발행 비용 = 쿠폰 수수료 합계 → ② 신규 유입 = 신규 사용자 수 → ③ 추정 매출 = 사용 횟수 × 평균 객단가",
        "ROI 카드 + \"쿠폰 1장 = 단골 N명\" 환산",
        "첫 주는 \"다음 주부터 ROI를 확인\"",
        "SCR-OWN-008", "API-GET-stores/{id}/roi?period=week"),
    "F-0507": ("쿠폰별 발행/사용/사용률/시간대별 사용 분포. 저성과 쿠폰엔 가이드 메시지.",
        "storeId, couponId (선택)",
        "① 쿠폰별 집계 → ② 사용률 < 20% 쿠폰엔 \"할인율 +10% 조정 권장\" 등 자동 가이드",
        "표 + 시간대별 히트맵 + 가이드 메시지",
        "쿠폰 0건 시 \"첫 쿠폰을 발행해보세요\" CTA",
        "SCR-OWN-009", "API-GET-stores/{id}/coupons/analytics"),
    "F-9001": ("휴대폰 본인 인증 또는 카카오·네이버 소셜 로그인. 소셜 시 본인 인증 별도 1회.",
        "가입 폼 또는 OAuth 콜백",
        "① 소셜 토큰 검증 → ② 사용자 생성 → ③ 본인 인증 (F-0302) → ④ 가입 완료",
        "JWT 액세스/리프레시 토큰",
        "중복 가입 시 \"이미 가입된 휴대폰입니다\"",
        "SCR-USR-011", "API-POST-auth/signup, API-POST-auth/oauth/callback"),
    "F-9002": ("F-9001 + 사업자 인증 (F-0401) 결합 흐름.",
        "폼", "F-9001 → F-0401 순차 진행",
        "사장님 계정 + 매장 등록 진입",
        "사업자 미인증 시 매장 등록 불가",
        "SCR-OWN-002", "API-POST-auth/signup-owner"),
    "F-9003": ("사용자/사장님/어드민 3종. 어드민은 계정 정지·복구 가능.",
        "JWT role 클레임",
        "미들웨어로 권한 검증, 403 또는 통과",
        "권한별 라우팅",
        "권한 미달 시 403 + 안내 페이지",
        "SCR-ADM-001", "모든 API에 권한 검증 미들웨어"),
    "F-9004": ("쿠폰 사용 거부·부정 사용 등 신고 접수 → 24시간 내 1차 응답.",
        "신고 폼 (사용자/사장님)",
        "① 신고 INSERT → ② 어드민 대시보드 알림 → ③ 24h SLA 모니터링",
        "신고 ID + 처리 상태",
        "동일 사용자 24h 내 중복 신고 시 통합",
        "SCR-USR-012, SCR-ADM-002", "API-POST-reports, API-PATCH-reports/{id}"),
    "F-9005": ("개인정보 즉시 파기, 거래 기록(쿠폰 사용 등) 5년 보관.",
        "탈퇴 사유 (선택)",
        "① 30일 유예 (복구 가능) → ② 30일 후 개인정보 컬럼 NULL/HASH → ③ 5년 후 거래 기록 익명화",
        "탈퇴 완료",
        "미사용 쿠폰 존재 시 사전 안내",
        "SCR-USR-013", "API-DELETE-users/me"),
    "F-9006": ("시스템 공지·이벤트·쿠폰 사용·만료 통합 알림 센터. 읽음/안 읽음 구분.",
        "userId",
        "① DB 조회 + 페이지네이션 → ② 읽음 처리 PATCH → ③ 30일 후 자동 정리 (크론)",
        "알림 리스트",
        "빈 알림함 → \"새 알림이 없어요\"",
        "SCR-USR-010", "API-GET-notifications, API-PATCH-notifications/{id}/read"),
    "F-9007": ("폰트 1.25x / 1.5x 확대. 50대 사장님·주부 페르소나 대응.",
        "fontScale: 1.0|1.25|1.5",
        "CSS root 변수 동적 변경 + LocalStorage 영속",
        "전체 화면 폰트 확대",
        "레이아웃 깨짐 방지 위해 컨테이너 max-width 별도",
        "SCR-USR-014, SCR-OWN-010", "(클라이언트)"),
    "F-9008": ("스크린리더 호환, alt 텍스트, 키보드 내비게이션, ARIA 라벨. WCAG 2.1 AA.",
        "모든 UI 컴포넌트",
        "① semantic 태그 → ② 모든 이미지 alt → ③ 대비 4.5:1 이상 → ④ tab 순서 정의",
        "접근성 준수 UI",
        "Lighthouse Accessibility 90+ 자동 검증 (CI)",
        "전체", "—"),
}

SHOULD_DETAIL = {
    "F-0508": ("월 단위 일괄 내보내기. 매장 회계용.", "SCR-OWN-008", "API-GET-stores/{id}/metrics/export?format=csv|xlsx&month="),
    "F-7001": ("쿠폰 사용 후 24시간 내 작성. 사장님 답글 1회. 매장 1,000+ 이후 활성화.", "SCR-USR-015", "API-POST-stores/{id}/reviews, API-POST-reviews/{id}/reply"),
    "F-7002": ("사용자가 매장 즐겨찾기 등록. 신규 특선 발행 시 알림.", "SCR-USR-016", "API-POST-users/me/favorites, API-GET-users/me/favorites"),
    "F-7003": ("11~14시 사용자 위치 500m 내 새 점심특선 발행 시 알림. 옵트인 + 일 3건 제한.", "SCR-USR-014", "API-POST-notifications/trigger-by-location (서버 내부)"),
    "F-7004": ("즐겨찾기 매장 신규 특선·쿠폰 발행 시 알림. 옵트인 + 매장당 일 1건 제한.", "SCR-USR-014", "(내부)"),
    "F-7005": ("HR 대시보드에서 임직원 등록 + 식권 일괄 발행. 84.6% 선호 정량 근거.", "SCR-OWN-011", "API-POST-b2b/{companyId}/employees/bulk, /meal-tickets/issue"),
    "F-7006": ("사용처별·임직원별 식대 정산. 월말 CSV/PDF 자동 생성.", "SCR-OWN-012", "API-GET-b2b/{companyId}/settlements?month="),
    "F-7007": ("점심 카테고리와 분리된 별도 탭. 시간대 자동 전환. DAU 한계 대응.", "SCR-USR-005", "API-GET-lunch-specials?timeSlot=dinner|weekend"),
}

COULD_DETAIL = {
    "F-8001": ("메인 상단 고정·카테고리 1위 슬롯·키워드 광고. 사장님 셀프 서비스 구매. 노출/클릭/전환 리포트.", "SCR-OWN-013", "API-POST-ads/campaigns, API-GET-ads/campaigns/{id}/report"),
    "F-8002": ("KRW zero-decimal, 결제 confirm API, 환불·취소 처리.", "SCR-USR-017", "API-POST-payments/intent, /confirm, /refund"),
    "F-8003": ("토스페이먼츠 통합 모듈 활용.", "SCR-USR-017", "F-8002와 동일 (paymentMethod 분기)"),
    "F-8004": ("카드 공유 → 카카오톡/링크. 추천인 보상 캠페인 연계.", "SCR-USR-018", "API-POST-invitations, API-GET-invitations/me"),
}

SCREENS_USR = [
    ("SCR-USR-001", "메인 지도", "F-0101~107"),
    ("SCR-USR-002", "메인 리스트", "F-0102"),
    ("SCR-USR-003", "카테고리 필터 모달", "F-0104"),
    ("SCR-USR-004", "검색", "F-0108"),
    ("SCR-USR-005", "메인 피드 (점심특선)", "F-0201~205, 207, F-7007"),
    ("SCR-USR-006", "점심특선 상세", "F-0206, F-0301"),
    ("SCR-USR-007", "내 쿠폰함", "F-0309, F-0307"),
    ("SCR-USR-008", "본인 인증", "F-0302"),
    ("SCR-USR-009", "쿠폰 QR 제시", "F-0303"),
    ("SCR-USR-010", "알림 센터", "F-9006, F-0308"),
    ("SCR-USR-011", "회원가입", "F-9001"),
    ("SCR-USR-012", "신고", "F-9004"),
    ("SCR-USR-013", "회원 탈퇴", "F-9005"),
    ("SCR-USR-014", "설정 (큰 글씨·알림)", "F-9007, F-7003, F-7004"),
    ("SCR-USR-015", "리뷰 작성/조회", "F-7001"),
    ("SCR-USR-016", "즐겨찾기", "F-7002"),
    ("SCR-USR-017", "결제 (Phase 3)", "F-8002/3"),
    ("SCR-USR-018", "친구 초대 (Phase 3)", "F-8004"),
]

SCREENS_OWN = [
    ("SCR-OWN-001", "QR 스캐너", "F-0304, F-0305"),
    ("SCR-OWN-002", "사업자 인증", "F-0401, F-9002"),
    ("SCR-OWN-003", "매장 등록 위저드 (3단계)", "F-0402, F-0403"),
    ("SCR-OWN-004", "점심특선 등록", "F-0404~406"),
    ("SCR-OWN-005", "홈 대시보드 (상태 토글)", "F-0407"),
    ("SCR-OWN-006", "매장 관리 (수정·이력)", "F-0408"),
    ("SCR-OWN-007", "쿠폰 발행", "F-0501~503"),
    ("SCR-OWN-008", "통계 대시보드 (4지표·ROI)", "F-0504~506, F-0508"),
    ("SCR-OWN-009", "쿠폰별 분석", "F-0507"),
    ("SCR-OWN-010", "설정 (큰 글씨)", "F-9007"),
    ("SCR-OWN-011", "B2B 대시보드 (Phase 2)", "F-7005"),
    ("SCR-OWN-012", "B2B 정산 (Phase 2)", "F-7006"),
    ("SCR-OWN-013", "광고 (Phase 3)", "F-8001"),
]

SCREENS_ADM = [
    ("SCR-ADM-001", "어드민 홈", "F-9003"),
    ("SCR-ADM-002", "신고 처리", "F-9004"),
]

APIS = [
    ("API-GET-restaurants", "GET", "/api/restaurants", "F-0101~108"),
    ("API-GET-search", "GET", "/api/search", "F-0108"),
    ("API-GET-lunch-specials", "GET", "/api/lunch-specials", "F-0201~207"),
    ("API-GET-lunch-specials-detail", "GET", "/api/lunch-specials/{id}", "F-0206"),
    ("API-POST-coupons-issue", "POST", "/api/coupons/issue", "F-0301"),
    ("API-POST-auth-verify-identity", "POST", "/api/auth/verify-identity", "F-0302"),
    ("API-POST-coupons-token", "POST", "/api/coupons/{id}/token", "F-0303"),
    ("API-POST-coupons-redeem", "POST", "/api/coupons/redeem", "F-0304, F-0305"),
    ("API-GET-coupons-me", "GET", "/api/coupons/me", "F-0307, F-0309"),
    ("API-GET-notifications", "GET", "/api/notifications", "F-0308, F-9006"),
    ("API-POST-owners-verify-business", "POST", "/api/owners/verify-business", "F-0401"),
    ("API-POST-stores", "POST", "/api/stores", "F-0402"),
    ("API-PATCH-stores-meta", "PATCH", "/api/stores/{id}/meta", "F-0403"),
    ("API-POST-lunch-specials", "POST", "/api/stores/{id}/lunch-specials", "F-0404"),
    ("API-POST-recurring-specials", "POST", "/api/stores/{id}/recurring-specials", "F-0405"),
    ("API-PATCH-lunch-specials-status", "PATCH", "/api/lunch-specials/{id}/status", "F-0406"),
    ("API-PATCH-stores-status", "PATCH", "/api/stores/{id}/status", "F-0407"),
    ("API-PATCH-stores", "PATCH", "/api/stores/{id}", "F-0408"),
    ("API-POST-coupons", "POST", "/api/stores/{id}/coupons", "F-0501~503"),
    ("API-GET-metrics", "GET", "/api/stores/{id}/metrics", "F-0504, F-0505"),
    ("API-GET-roi", "GET", "/api/stores/{id}/roi", "F-0506"),
    ("API-GET-coupons-analytics", "GET", "/api/stores/{id}/coupons/analytics", "F-0507"),
    ("API-POST-auth-signup", "POST", "/api/auth/signup", "F-9001"),
    ("API-POST-auth-signup-owner", "POST", "/api/auth/signup-owner", "F-9002"),
    ("API-POST-reports", "POST", "/api/reports", "F-9004"),
    ("API-DELETE-users-me", "DELETE", "/api/users/me", "F-9005"),
    ("API-GET-metrics-export", "GET", "/api/stores/{id}/metrics/export", "F-0508 (Phase 2)"),
    ("API-POST-reviews", "POST", "/api/stores/{id}/reviews", "F-7001 (Phase 2)"),
    ("API-POST-favorites", "POST", "/api/users/me/favorites", "F-7002 (Phase 2)"),
    ("API-POST-b2b-employees", "POST", "/api/b2b/{companyId}/employees/bulk", "F-7005 (Phase 2)"),
    ("API-GET-b2b-settlements", "GET", "/api/b2b/{companyId}/settlements", "F-7006 (Phase 2)"),
    ("API-POST-ads", "POST", "/api/ads/campaigns", "F-8001 (Phase 3)"),
    ("API-POST-payments", "POST", "/api/payments/intent, /confirm, /refund", "F-8002/3 (Phase 3)"),
    ("API-POST-invitations", "POST", "/api/invitations", "F-8004 (Phase 3)"),
]

# ============ Workbook 생성 ============
wb = Workbook()
wb.remove(wb.active)

# ===== Sheet 1: 기능 목록 (전체 59건) =====
ws1 = wb.create_sheet("01.기능 목록")
add_title(ws1, "점심특강 기능 명세서 — Master List (총 59건)", 5)
headers = ["F-ID", "기능명", "요구사항 ID", "우선순위", "단계"]
ws1.append([])  # blank row 2
ws1.append(headers)
style_header_row(ws1, 3)

row = 4
for item in MUST_LIST + SHOULD_LIST + COULD_LIST:
    ws1.append(item)
    # 우선순위 색상
    prio_cell = ws1.cell(row=row, column=4)
    prio_cell.fill = priority_fill(item[3])
    prio_cell.font = FONT_BOLD
    prio_cell.alignment = ALIGN_CENTER
    row += 1

style_data_cells(ws1, 4, row - 1)
set_col_widths(ws1, [12, 38, 16, 12, 14])
ws1.freeze_panes = "A4"

# ===== Sheet 2: F1-F5 + OPS 풀 명세 =====
ws2 = wb.create_sheet("02.MVP 기능 풀 명세")
add_title(ws2, "MVP Must 47건 풀 명세 (7요소)", 8)
ws2.append([])
headers2 = ["F-ID", "기능명", "설명", "입력", "처리", "출력", "예외 처리", "관련 화면 / API"]
ws2.append(headers2)
style_header_row(ws2, 3)

row = 4
for fid, name, _, _, _ in MUST_LIST:
    if fid in DETAILED:
        d = DETAILED[fid]
        screen_api = f"{d[5]}\n{d[6]}"
        ws2.append([fid, name, d[0], d[1], d[2], d[3], d[4], screen_api])
        # F-ID 셀 하이라이트
        ws2.cell(row=row, column=1).font = FONT_BOLD
        ws2.cell(row=row, column=1).fill = FILL_GRAY_LIGHT
        ws2.cell(row=row, column=1).alignment = ALIGN_CENTER
        row += 1

style_data_cells(ws2, 4, row - 1)
set_col_widths(ws2, [10, 26, 38, 32, 42, 28, 36, 28])
ws2.freeze_panes = "C4"
# 행 높이 자동
for r in range(4, row):
    ws2.row_dimensions[r].height = 95

# ===== Sheet 3: Phase 2-3 약식 명세 =====
ws3 = wb.create_sheet("03.Phase 2-3 약식")
add_title(ws3, "Phase 2 (Should · M4+) & Phase 3 (Could · M7+) 약식 명세 12건", 5)
ws3.append([])
ws3.append(["F-ID", "기능명", "설명 (약식)", "관련 화면", "관련 API"])
style_header_row(ws3, 3)

row = 4
for fid, name, _, prio, _ in SHOULD_LIST:
    if fid in SHOULD_DETAIL:
        d = SHOULD_DETAIL[fid]
        ws3.append([fid, name, d[0], d[1], d[2]])
        ws3.cell(row=row, column=1).font = FONT_BOLD
        ws3.cell(row=row, column=1).fill = FILL_SHOULD
        ws3.cell(row=row, column=1).alignment = ALIGN_CENTER
        row += 1
for fid, name, _, prio, _ in COULD_LIST:
    if fid in COULD_DETAIL:
        d = COULD_DETAIL[fid]
        ws3.append([fid, name, d[0], d[1], d[2]])
        ws3.cell(row=row, column=1).font = FONT_BOLD
        ws3.cell(row=row, column=1).fill = FILL_COULD
        ws3.cell(row=row, column=1).alignment = ALIGN_CENTER
        row += 1

style_data_cells(ws3, 4, row - 1)
set_col_widths(ws3, [10, 28, 60, 22, 50])
ws3.freeze_panes = "A4"
for r in range(4, row):
    ws3.row_dimensions[r].height = 50

# ===== Sheet 4: 화면 ID 총괄 =====
ws4 = wb.create_sheet("04.화면 ID (33종)")
add_title(ws4, "화면 ID 사전 정의 — 사용자 18 + 사장님 13 + 어드민 2 = 33종", 4)
ws4.append([])
ws4.append(["구분", "화면 ID", "화면명", "관련 기능"])
style_header_row(ws4, 3)

row = 4
def fill_section_for_screens(ws, label, items, fill, start_row):
    r = start_row
    for sid, name, fids in items:
        ws.append([label, sid, name, fids])
        ws.cell(row=r, column=1).font = FONT_BOLD
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2).font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        r += 1
    return r

row = fill_section_for_screens(ws4, "사용자 (USR)", SCREENS_USR, FILL_MUST, row)
row = fill_section_for_screens(ws4, "사장님 (OWN)", SCREENS_OWN, FILL_SHOULD, row)
row = fill_section_for_screens(ws4, "어드민 (ADM)", SCREENS_ADM, FILL_COULD, row)

style_data_cells(ws4, 4, row - 1)
set_col_widths(ws4, [16, 18, 32, 38])
ws4.freeze_panes = "A4"

# ===== Sheet 5: API 엔드포인트 =====
ws5 = wb.create_sheet("05.API 엔드포인트 (34종)")
add_title(ws5, "API 엔드포인트 사전 정의 — 34종 (MVP + Phase 2/3)", 4)
ws5.append([])
ws5.append(["API ID", "METHOD", "PATH", "관련 기능"])
style_header_row(ws5, 3)

method_fills = {
    "GET": PatternFill("solid", start_color="E3F2FD"),
    "POST": PatternFill("solid", start_color="E8F5E9"),
    "PATCH": PatternFill("solid", start_color="FFF3E0"),
    "DELETE": PatternFill("solid", start_color="FFE0E0"),
}
row = 4
for aid, method, path, fids in APIS:
    ws5.append([aid, method, path, fids])
    ws5.cell(row=row, column=2).fill = method_fills.get(method, FILL_WHITE)
    ws5.cell(row=row, column=2).font = FONT_BOLD
    ws5.cell(row=row, column=2).alignment = ALIGN_CENTER
    ws5.cell(row=row, column=3).font = Font(name="Consolas", size=10, color="111827")
    row += 1

style_data_cells(ws5, 4, row - 1)
set_col_widths(ws5, [32, 10, 42, 28])
ws5.freeze_panes = "A4"

# ===== Sheet 6: 요약 통계 (수식 사용) =====
ws6 = wb.create_sheet("06.요약 통계")
add_title(ws6, "기능 명세 요약 통계 (자동 집계)", 4)
ws6.append([])

# 우선순위 집계 (수식)
ws6.append(["우선순위별 기능 수"])
ws6.cell(row=3, column=1).font = FONT_BOLD
ws6.cell(row=3, column=1).fill = FILL_NAVY
ws6.cell(row=3, column=1).font = Font(name="Malgun Gothic", size=12, bold=True, color="FFFFFF")
ws6.merge_cells(start_row=3, end_row=3, start_column=1, end_column=4)
ws6.cell(3, 1).alignment = ALIGN_CENTER

ws6.append(["우선순위", "기능 수", "비율", "비고"])
style_header_row(ws6, 4)

ws6.append(["Must (P0)",   "=COUNTIF('01.기능 목록'!D:D,\"Must\")",   "=B5/$B$8", "MVP 핵심"])
ws6.append(["Should (P1)", "=COUNTIF('01.기능 목록'!D:D,\"Should\")", "=B6/$B$8", "Phase 2"])
ws6.append(["Could (P2)",  "=COUNTIF('01.기능 목록'!D:D,\"Could\")",  "=B7/$B$8", "Phase 3"])
ws6.append(["합계",         "=SUM(B5:B7)",                              "=B8/$B$8", "전체"])

for r in [5, 6, 7]:
    ws6.cell(row=r, column=1).font = FONT_BOLD
    ws6.cell(row=r, column=1).fill = priority_fill(ws6.cell(row=r, column=1).value.split()[0])
    ws6.cell(row=r, column=1).alignment = ALIGN_CENTER
    ws6.cell(row=r, column=3).number_format = "0.0%"
ws6.cell(row=8, column=1).font = FONT_BOLD
ws6.cell(row=8, column=1).fill = FILL_NAVY
ws6.cell(row=8, column=1).font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
ws6.cell(row=8, column=1).alignment = ALIGN_CENTER
ws6.cell(row=8, column=3).number_format = "0.0%"
style_data_cells(ws6, 5, 8)

# 단계별 집계
ws6.append([])
ws6.append(["단계별 기능 수 (출시 마일스톤)"])
ws6.cell(row=10, column=1).font = Font(name="Malgun Gothic", size=12, bold=True, color="FFFFFF")
ws6.cell(row=10, column=1).fill = FILL_NAVY
ws6.merge_cells(start_row=10, end_row=10, start_column=1, end_column=4)
ws6.cell(10, 1).alignment = ALIGN_CENTER

ws6.append(["단계", "기능 수", "비율", "비고"])
style_header_row(ws6, 11)
ws6.append(["M2 알파",  "=COUNTIF('01.기능 목록'!E:E,\"M2 알파\")", "=B12/SUM($B$12:$B$15)", "W6-W12 (F1+F4 일부)"])
ws6.append(["M3 베타",  "=COUNTIF('01.기능 목록'!E:E,\"M3 베타\")", "=B13/SUM($B$12:$B$15)", "W13-W22"])
ws6.append(["M4+",      "=COUNTIF('01.기능 목록'!E:E,\"M4+\")",     "=B14/SUM($B$12:$B$15)", "Phase 2"])
ws6.append(["M7+",      "=COUNTIF('01.기능 목록'!E:E,\"M7+\")",     "=B15/SUM($B$12:$B$15)", "Phase 3"])
for r in [12, 13, 14, 15]:
    ws6.cell(row=r, column=3).number_format = "0.0%"
    ws6.cell(row=r, column=1).font = FONT_BOLD
style_data_cells(ws6, 12, 15)

# 자산 카운트
ws6.append([])
ws6.append(["사전 정의 자산 (다음 산출물 입력값)"])
ws6.cell(row=17, column=1).font = Font(name="Malgun Gothic", size=12, bold=True, color="FFFFFF")
ws6.cell(row=17, column=1).fill = FILL_NAVY
ws6.merge_cells(start_row=17, end_row=17, start_column=1, end_column=4)
ws6.cell(17, 1).alignment = ALIGN_CENTER

ws6.append(["자산", "개수 (수식)", "다음 산출물", "비고"])
style_header_row(ws6, 18)
ws6.append(["화면 ID (전체)",      "=COUNTA('04.화면 ID (33종)'!B:B)-2", "#8 정보구조도 + #9 화면설계서", "USR/OWN/ADM 합산 (헤더 제외)"])
ws6.append(["API 엔드포인트",       "=COUNTA('05.API 엔드포인트 (34종)'!A:A)-2", "#7 API스펙",         "OpenAPI 3.0 확장"])
for r in [19, 20]:
    ws6.cell(row=r, column=1).font = FONT_BOLD
style_data_cells(ws6, 19, 20)

set_col_widths(ws6, [22, 18, 38, 28])

# 문서 메타
ws6.append([])
ws6.append(["문서 메타"])
ws6.cell(row=22, column=1).font = Font(name="Malgun Gothic", size=12, bold=True, color="FFFFFF")
ws6.cell(row=22, column=1).fill = FILL_NAVY
ws6.merge_cells(start_row=22, end_row=22, start_column=1, end_column=4)
ws6.cell(22, 1).alignment = ALIGN_CENTER

meta = [
    ("프로젝트명", "점심특강 (Lunch Special Lecture)"),
    ("원본 문서", "02.기획문서/기능명세서.md v1.0"),
    ("작성일", "2026-05-31"),
    ("작성자", "PM (점심특강 기획팀)"),
    ("근거 문서", "요구사항정의서 v1.0 · 서비스기획서 v1.0 · 마켓리서치 v1.1"),
]
for i, (k, v) in enumerate(meta):
    r = 23 + i
    ws6.cell(row=r, column=1, value=k).font = FONT_BOLD
    ws6.cell(row=r, column=2, value=v)
    ws6.merge_cells(start_row=r, end_row=r, start_column=2, end_column=4)
style_data_cells(ws6, 23, 22 + len(meta))

wb.save(OUT)
print(f"✓ Saved: {OUT}")
